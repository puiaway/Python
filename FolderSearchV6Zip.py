import os
import csv
import tempfile
import threading
import platform
import chardet
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
import json
import queue
import zipfile
import io
import shutil

# --- Optional Dependency for Excel ---
try:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# --- Optional Dependency for 7z ---
try:
    import py7zr
    PY7ZR_AVAILABLE = True
except ImportError:
    PY7ZR_AVAILABLE = False

# --- Constants ---
HISTORY_FILE = os.path.expanduser("~/.text_search_app_history.json")
DEFAULT_EXTENSIONS = ".txt, .log, .csv, .json, .xml, .md, .py"
ARCHIVE_EXTS = ('.zip', '.7z')

# --- Helper Functions ---
def safe_path(path):
    """Return a Windows-safe extended-length path when needed, including UNC."""
    if platform.system() != 'Windows':
        return path
    path = os.path.abspath(path)
    if path.startswith('\\\\?\\'):
        return path
    if path.startswith('\\\\'):  # UNC path
        return '\\\\?\\UNC\\' + path.lstrip('\\')
    return '\\\\?\\' + path

def detect_encoding(file_path, sample_size=10000):
    """Detects file encoding using a sample of the file (long-path safe)."""
    fp = safe_path(file_path)
    with open(fp, 'rb') as f:
        raw_data = f.read(sample_size)
    return chardet.detect(raw_data)['encoding'] or 'utf-8'


class TextSearchApp:
    def __init__(self, master):
        self.master = master
        master.title("Text Search Pro")
        master.geometry("980x740")
        master.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.stop_flag = threading.Event()
        self.temp_csv_path = None

        # Thread-safe UI queue
        self.ui_queue = queue.Queue()
        self.master.after(100, self._drain_ui_queue)

        self._create_widgets()
        self.keyword_history = self._load_keyword_history()
        self.keyword_entry['values'] = self.keyword_history

    # ---------------- UI scaffolding ----------------
    def _create_widgets(self):
        main_frame = ttk.Frame(self.master, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew")
        self.master.rowconfigure(0, weight=1)
        self.master.columnconfigure(0, weight=1)

        # Input widgets
        ttk.Label(main_frame, text="Folder:").grid(row=0, column=0, sticky='w', padx=5, pady=2)
        self.folder_entry = ttk.Entry(main_frame, width=82)
        self.folder_entry.grid(row=0, column=1, columnspan=2, sticky="ew", padx=5, pady=2)
        ttk.Button(main_frame, text="Browse...", command=self.browse_folder).grid(row=0, column=3, padx=5)

        ttk.Label(main_frame, text="Keyword:").grid(row=1, column=0, sticky='w', padx=5, pady=2)
        self.keyword_entry = ttk.Combobox(main_frame, width=80)
        self.keyword_entry.grid(row=1, column=1, columnspan=2, sticky="ew", padx=5, pady=2)

        ttk.Label(main_frame, text="File Types:").grid(row=2, column=0, sticky='w', padx=5, pady=2)
        self.extensions_entry = ttk.Entry(main_frame, width=80)
        self.extensions_entry.grid(row=2, column=1, columnspan=2, sticky="ew", padx=5, pady=2)
        self.extensions_entry.insert(0, DEFAULT_EXTENSIONS)

        self.search_btn = ttk.Button(main_frame, text="Search", command=self.start_search_thread)
        self.search_btn.grid(row=1, column=3, padx=5)

        # Options
        options_frame = ttk.LabelFrame(main_frame, text="Options", padding="10")
        options_frame.grid(row=3, column=0, columnspan=4, sticky='ew', padx=5, pady=5)

        self.match_once_var = tk.BooleanVar()
        self.last_match_var = tk.BooleanVar()
        self.case_sensitive_var = tk.BooleanVar()
        self.show_all_var = tk.BooleanVar()
        self.include_nomatch_var = tk.BooleanVar(value=True)
        self.search_archives_var = tk.BooleanVar(value=True)

        ttk.Checkbutton(options_frame, text="Match once per file", variable=self.match_once_var).grid(row=0, column=0, sticky='w', padx=5)
        ttk.Checkbutton(options_frame, text="Use last match", variable=self.last_match_var).grid(row=0, column=1, sticky='w', padx=5)
        ttk.Checkbutton(options_frame, text="Case sensitive", variable=self.case_sensitive_var).grid(row=0, column=2, sticky='w', padx=5)
        ttk.Checkbutton(options_frame, text="Include non-matching files", variable=self.include_nomatch_var).grid(row=0, column=3, sticky='w', padx=5)
        ttk.Checkbutton(options_frame, text="Show all in preview", variable=self.show_all_var).grid(row=0, column=4, sticky='w', padx=5)
        ttk.Checkbutton(options_frame, text="Search inside archives (.zip, .7z)", variable=self.search_archives_var).grid(row=0, column=5, sticky='w', padx=5)

        # Progress & Status
        self.progress = ttk.Progressbar(main_frame, orient="horizontal", mode="determinate")
        self.progress.grid(row=4, column=0, columnspan=4, sticky="ew", padx=5, pady=5)
        self.status_label = ttk.Label(main_frame, text="Status: Ready")
        self.status_label.grid(row=5, column=0, columnspan=2, sticky="w", padx=5)

        # Action Buttons
        action_frame = ttk.Frame(main_frame)
        action_frame.grid(row=5, column=2, columnspan=2, sticky="e", padx=5)
        self.cancel_btn = ttk.Button(action_frame, text="Cancel", command=self.cancel_search, state=tk.DISABLED)
        self.cancel_btn.pack(side=tk.LEFT, padx=5)
        self.export_csv_btn = ttk.Button(action_frame, text="Export CSV", command=self.export_csv, state=tk.DISABLED)
        self.export_csv_btn.pack(side=tk.LEFT, padx=5)
        self.export_excel_btn = ttk.Button(action_frame, text="Export Excel", command=self.export_excel, state=tk.DISABLED)
        self.export_excel_btn.pack(side=tk.LEFT, padx=5)
        if not OPENPYXL_AVAILABLE:
            self.export_excel_btn.config(text="Excel (install openpyxl)")

        # Result Area
        self.result_area = scrolledtext.ScrolledText(main_frame, width=120, height=25, wrap=tk.WORD)
        self.result_area.grid(row=6, column=0, columnspan=4, sticky="nsew", padx=5, pady=5)

        main_frame.rowconfigure(6, weight=1)
        main_frame.columnconfigure(1, weight=1)

    def _load_keyword_history(self):
        if not os.path.exists(HISTORY_FILE):
            return []
        try:
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return []

    def _save_keyword_history(self):
        try:
            with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.keyword_history, f, indent=4)
        except IOError:
            pass  # Silently fail

    def browse_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.folder_entry.delete(0, tk.END)
            self.folder_entry.insert(0, folder)

    def _set_ui_state(self, searching=False):
        """Enable/disable UI elements and export buttons based on state & data."""
        state = tk.DISABLED if searching else tk.NORMAL
        for w in (self.search_btn, self.folder_entry, self.keyword_entry, self.extensions_entry):
            w.config(state=state)

        self.cancel_btn.config(state=(tk.NORMAL if searching else tk.DISABLED))

        has_data = (not searching) and self.temp_csv_path and os.path.exists(self.temp_csv_path)
        self.export_csv_btn.config(state=(tk.NORMAL if has_data else tk.DISABLED))
        self.export_excel_btn.config(
            state=(tk.NORMAL if has_data and OPENPYXL_AVAILABLE else tk.DISABLED)
        )
        if not OPENPYXL_AVAILABLE:
            self.export_excel_btn.config(text="Excel (install openpyxl)")

    # ---------------- Thread helpers ----------------
    def _post_ui(self, fn, *args, **kwargs):
        """Schedule a UI update from worker threads."""
        self.ui_queue.put((fn, args, kwargs))

    def _drain_ui_queue(self):
        try:
            while True:
                fn, args, kwargs = self.ui_queue.get_nowait()
                try:
                    fn(*args, **kwargs)
                except Exception:
                    # Avoid crashing UI loop due to a bad callback
                    pass
        except queue.Empty:
            pass
        # Poll frequently to keep UI responsive
        self.master.after(50, self._drain_ui_queue)

    # ---------------- Search flow ----------------
    def start_search_thread(self):
        """Validates input and starts the search in a new thread."""
        folder = self.folder_entry.get()
        keyword = self.keyword_entry.get().strip()
        extensions = self.extensions_entry.get().strip()

        if not os.path.isdir(folder) or not keyword or not extensions:
            messagebox.showerror("Error", "Please provide a valid folder, keyword, and file types.")
            return

        # Update and save keyword history
        if keyword not in self.keyword_history:
            self.keyword_history.insert(0, keyword)
            self.keyword_history = self.keyword_history[:20]
            self.keyword_entry['values'] = self.keyword_history
            self._save_keyword_history()

        self.stop_flag.clear()
        self._set_ui_state(searching=True)
        self.result_area.delete(1.0, tk.END)
        self.status_label.config(text="Status: Preparing to search...")
        self.progress["value"] = 0

        search_thread = threading.Thread(
            target=self.search_files,
            args=(folder, keyword, extensions),
            daemon=True
        )
        search_thread.start()

    def cancel_search(self):
        self.status_label.config(text="Status: Cancelling...")
        self.stop_flag.set()

    def search_files(self, folder, keyword, extensions_str):
        """The core search logic that runs in a background thread."""
        # Prepare params
        params = {
            "match_once": self.match_once_var.get(),
            "use_last": self.last_match_var.get(),
            "case_sens": self.case_sensitive_var.get(),
            "show_all": self.show_all_var.get(),
            "inc_nomatch": self.include_nomatch_var.get()
        }
        keyword_to_find = keyword if params["case_sens"] else keyword.lower()

        # Normalize extension filters
        raw_exts = [e.strip().lower() for e in extensions_str.split(',') if e.strip()]
        file_exts = tuple(e if e.startswith('.') else f'.{e}' for e in raw_exts)

        # Build file and archive lists
        normal_files = []
        archive_files = []
        for root, _, files in os.walk(folder):
            for name in files:
                lower = name.lower()
                full = os.path.join(root, name)
                if lower.endswith(ARCHIVE_EXTS) and self.search_archives_var.get():
                    archive_files.append(full)
                elif lower.endswith(file_exts):
                    normal_files.append(full)

        file_list = normal_files + archive_files
        total_files = len(file_list)
        if total_files == 0:
            self._post_ui(self.status_label.config, text="Status: No matching file types found.")
            self._post_ui(self._set_ui_state, False)
            return

        self._post_ui(self.progress.config, maximum=total_files)

        # Create temporary CSV
        try:
            temp_file = tempfile.NamedTemporaryFile(delete=False, mode='w', encoding='utf-8', newline='')
            self.temp_csv_path = temp_file.name
            writer = csv.writer(temp_file)
            writer.writerow(["Path", "Filename", "Line Number", "Line Content"])
        except IOError as e:
            self._post_ui(messagebox.showerror, "File Error", f"Could not create a temporary file: {e}")
            self._post_ui(self._set_ui_state, False)
            return

        # Main processing
        preview_results, total_matches = self._process_files(
            file_list=file_list,
            writer=writer,
            keyword=keyword_to_find,
            params=params,
            file_exts=file_exts,
            base_folder=folder
        )

        # Finalize
        try:
            temp_file.close()
        except Exception:
            pass

        if self.stop_flag.is_set():
            msg = f"Status: Search cancelled. Processed {self.progress['value']} files."
        else:
            msg = f"Status: Complete. Found {total_matches} matches in {total_files} files."

        self._post_ui(self.status_label.config, text=msg)
        self._post_ui(self.show_preview, preview_results, total_matches)
        self._post_ui(self._set_ui_state, False)

    # ---------------- File & archive processing ----------------
    def _iter_zip_member_lines(self, zip_path, info, encoding_hint=None):
        """Yield decoded text lines from a zip member."""
        with zipfile.ZipFile(safe_path(zip_path)) as zf:
            # First: detect encoding
            with zf.open(info, 'r') as f:
                head = f.read(10000)
            enc = chardet.detect(head)['encoding'] or encoding_hint or 'utf-8'
            # Then stream lines
            with zf.open(info, 'r') as f:
                for bline in f:
                    yield bline.decode(enc, errors='ignore')

    def _iter_7z_member_lines(self, sevenz_path, member_name, encoding_hint=None):
        """Yield decoded text lines from a 7z member using py7zr (in-memory BytesIO)."""
        if not PY7ZR_AVAILABLE:
            raise RuntimeError("py7zr not installed. Run: pip install py7zr")
        with py7zr.SevenZipFile(safe_path(sevenz_path), mode='r') as z:
            data_map = z.read([member_name])
            bio = data_map.get(member_name)
            if bio is None:
                return
            # detect encoding
            pos = bio.tell()
            head = bio.read(10000)
            enc = chardet.detect(head)['encoding'] or encoding_hint or 'utf-8'
            bio.seek(0)
            for bline in bio:
                yield bline.decode(enc, errors='ignore')

    def _process_archive(self, archive_path, writer, keyword, params, file_exts):
        """
        Scan an archive; return (preview_rows, match_count).
        Path column will be 'archive_path::member_path'.
        """
        preview_rows = []
        match_count = 0

        def maybe_preview(row):
            if params["show_all"] or len(preview_rows) < 1000:
                preview_rows.append(row)

        try:
            lower = archive_path.lower()

            if lower.endswith('.zip'):
                with zipfile.ZipFile(safe_path(archive_path)) as zf:
                    for info in zf.infolist():
                        if info.is_dir():
                            continue
                        inner = info.filename
                        if not inner.lower().endswith(file_exts):
                            continue

                        last_match_data = None
                        matches_in_member = []
                        line_no = 0
                        try:
                            for line in self._iter_zip_member_lines(archive_path, info):
                                if self.stop_flag.is_set():
                                    break
                                line_no += 1
                                line_to_check = line if params["case_sens"] else line.lower()
                                if keyword in line_to_check:
                                    match_data = (f"{archive_path}::{inner}", os.path.basename(inner), line_no, line.rstrip('\n'))
                                    if params["match_once"]:
                                        matches_in_member = [match_data]
                                        break
                                    elif params["use_last"]:
                                        last_match_data = match_data
                                    else:
                                        matches_in_member.append(match_data)
                            if params["use_last"] and last_match_data:
                                matches_in_member = [last_match_data]
                        except Exception as e:
                            matches_in_member = [(f"{archive_path}::{inner}", os.path.basename(inner), "ERROR", f"Could not read member: {e}")]

                        if matches_in_member or (params["inc_nomatch"] and not matches_in_member):
                            if not matches_in_member and params["inc_nomatch"]:
                                matches_in_member = [(f"{archive_path}::{inner}", os.path.basename(inner), "-", "[No match found]")]
                            for row in matches_in_member:
                                writer.writerow(row)
                                maybe_preview(row)
                            # Count true matches (exclude "-")
                            match_count += sum(1 for r in matches_in_member if isinstance(r[2], int) or (isinstance(r[2], str) and r[2].isdigit()))

            elif lower.endswith('.7z'):
                if not PY7ZR_AVAILABLE:
                    row = (archive_path, os.path.basename(archive_path), "ERROR", "Install py7zr to scan .7z (pip install py7zr)")
                    writer.writerow(row)
                    maybe_preview(row)
                    return preview_rows, match_count

                with py7zr.SevenZipFile(safe_path(archive_path), mode='r') as z:
                    for inner in z.getnames():
                        if inner.endswith('/') or inner.endswith('\\'):
                            continue
                        if not inner.lower().endswith(file_exts):
                            continue

                        last_match_data = None
                        matches_in_member = []
                        line_no = 0
                        try:
                            for line in self._iter_7z_member_lines(archive_path, inner):
                                if self.stop_flag.is_set():
                                    break
                                line_no += 1
                                line_to_check = line if params["case_sens"] else line.lower()
                                if keyword in line_to_check:
                                    match_data = (f"{archive_path}::{inner}", os.path.basename(inner), line_no, line.rstrip('\n'))
                                    if params["match_once"]:
                                        matches_in_member = [match_data]
                                        break
                                    elif params["use_last"]:
                                        last_match_data = match_data
                                    else:
                                        matches_in_member.append(match_data)
                            if params["use_last"] and last_match_data:
                                matches_in_member = [last_match_data]
                        except Exception as e:
                            matches_in_member = [(f"{archive_path}::{inner}", os.path.basename(inner), "ERROR", f"Could not read member: {e}")]

                        if matches_in_member or (params["inc_nomatch"] and not matches_in_member):
                            if not matches_in_member and params["inc_nomatch"]:
                                matches_in_member = [(f"{archive_path}::{inner}", os.path.basename(inner), "-", "[No match found]")]
                            for row in matches_in_member:
                                writer.writerow(row)
                                maybe_preview(row)
                            match_count += sum(1 for r in matches_in_member if isinstance(r[2], int) or (isinstance(r[2], str) and r[2].isdigit()))

            # else: not an archive (no-op)
        except Exception as e:
            row = (archive_path, os.path.basename(archive_path), "ERROR", f"Could not open archive: {e}")
            writer.writerow(row)
            maybe_preview(row)

        return preview_rows, match_count

    def _process_files(self, file_list, writer, keyword, params, file_exts, base_folder):
        """Iterates through files and archives and performs the search."""
        preview = []
        match_count = 0
        total_files = len(file_list)

        for i, file_path in enumerate(file_list, 1):
            if self.stop_flag.is_set():
                break

            lower_path = file_path.lower()
            is_archive = lower_path.endswith(ARCHIVE_EXTS)

            if is_archive and self.search_archives_var.get():
                arch_preview, arch_matches = self._process_archive(file_path, writer, keyword, params, file_exts)
                # Add to preview respecting the cap unless show_all
                if params["show_all"]:
                    preview.extend(arch_preview)
                else:
                    space = max(0, 1000 - len(preview))
                    if space > 0:
                        preview.extend(arch_preview[:space])
                match_count += arch_matches

            else:
                filename = os.path.basename(file_path)
                try:
                    rel_path = os.path.relpath(file_path, start=base_folder)
                except Exception:
                    rel_path = os.path.abspath(file_path)

                matches_in_file = []
                found_in_file = False

                try:
                    encoding = detect_encoding(file_path)
                    with open(safe_path(file_path), 'r', encoding=encoding, errors='ignore') as f:
                        last_match_data = None
                        for line_num, line in enumerate(f, 1):
                            if self.stop_flag.is_set():
                                break
                            line_to_check = line if params["case_sens"] else line.lower()
                            if keyword in line_to_check:
                                found_in_file = True
                                match_data = (rel_path, filename, line_num, line.rstrip('\n'))
                                if params["match_once"]:
                                    matches_in_file = [match_data]
                                    break
                                elif params["use_last"]:
                                    last_match_data = match_data
                                else:
                                    matches_in_file.append(match_data)
                        if params["use_last"] and last_match_data:
                            matches_in_file = [last_match_data]
                except Exception as e:
                    matches_in_file.append((rel_path, filename, "ERROR", f"Could not read file: {e}"))

                if found_in_file or (params["inc_nomatch"] and not matches_in_file):
                    if not matches_in_file and params["inc_nomatch"]:
                        matches_in_file.append((rel_path, filename, "-", "[No match found]"))
                    for match in matches_in_file:
                        writer.writerow(match)
                        if params["show_all"] or len(preview) < 1000:
                            preview.append(match)
                    match_count += len(matches_in_file) if found_in_file else 0

            # Progress updates (via UI thread)
            if i % 10 == 0 or i == total_files:
                self._post_ui(self.progress.config, value=i)
                self._post_ui(self.status_label.config, text=f"Status: Scanning {i}/{total_files} items...")

        return preview, match_count

    # ---------------- Preview & export ----------------
    def show_preview(self, preview, total_count):
        self.result_area.delete(1.0, tk.END)
        if not preview and not self.include_nomatch_var.get():
            self.result_area.insert(tk.END, "No matches found.")
            return

        for path_col, file_col, line_num, line in preview:
            self.result_area.insert(tk.END, f"{path_col} (Line {line_num}): {line}\n")

        if not self.show_all_var.get():
            # Count only true matches visible vs total matches
            shown_matches = sum(1 for _, _, ln, _ in preview if isinstance(ln, int) or (isinstance(ln, str) and ln.isdigit()))
            if total_count > shown_matches:
                self.result_area.insert(tk.END, f"\n--- Showing first {shown_matches} of {total_count} total matches ---\n")
                self.result_area.insert(tk.END, "Use 'Export CSV' or 'Export Excel' to view all results.")
        self.result_area.yview_moveto(0)

    def export_csv(self):
        if not self.temp_csv_path or not os.path.exists(self.temp_csv_path):
            messagebox.showwarning("No Data", "No results available to export.")
            return
        save_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv")],
            initialfile="search_results.csv"
        )
        if save_path:
            try:
                shutil.copy(self.temp_csv_path, save_path)
                messagebox.showinfo("Success", f"Results exported to\n{save_path}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export CSV:\n{e}")

    def export_excel(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Dependency Missing", "Please install the 'openpyxl' library to use this feature.\n\nCommand: pip install openpyxl")
            return

        if not self.temp_csv_path or not os.path.exists(self.temp_csv_path):
            messagebox.showwarning("No Data", "No results available to export.")
            return

        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile="search_results.xlsx"
        )
        if save_path:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Search Results"

                with open(self.temp_csv_path, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    for r_idx, row_data in enumerate(reader, 1):
                        for c_idx, cell_value in enumerate(row_data, 1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=cell_value)
                            if r_idx == 1:
                                cell.font = Font(bold=True)

                # Auto-fit column widths safely
                for col_idx in range(1, ws.max_column + 1):
                    max_len = 0
                    for row_idx in range(1, ws.max_row + 1):
                        val = ws.cell(row=row_idx, column=col_idx).value
                        if val is not None:
                            max_len = max(max_len, len(str(val)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 70)

                wb.save(save_path)
                messagebox.showinfo("Success", f"Results exported to\n{save_path}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to export to Excel:\n{e}")

    # ---------------- Cleanup ----------------
    def on_closing(self):
        """Handles window close event to clean up temporary files."""
        if self.temp_csv_path and os.path.exists(self.temp_csv_path):
            try:
                os.remove(self.temp_csv_path)
            except OSError:
                pass
        self.master.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app = TextSearchApp(root)
    root.mainloop()
